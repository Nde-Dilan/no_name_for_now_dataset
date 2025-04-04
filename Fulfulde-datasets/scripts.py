import pandas as pd
import openpyxl


def traduire_mots(fichier_doc1, fichier_doc2,
                  colonne_sortie):
    doc1 = pd.read_excel(fichier_doc1)
    doc2 = pd.read_excel(fichier_doc2)
    col_mot_doc1 = doc1.columns[1]
    col_mot_doc2 = doc2.columns[1]
    col_traduction_doc2 = doc2.columns[0]
    wb1 = openpyxl.load_workbook(fichier_doc1)
    ws1 = wb1.active
    wb2 = openpyxl.load_workbook(fichier_doc2)
    ws2 = wb2.active
    merged_cells_doc1 = ws1.merged_cells.ranges
    merged_cells_doc2 = ws2.merged_cells.ranges
    merged_rows_doc1 = []
    for merged_range in merged_cells_doc1:
        if merged_range.min_col <= 1 <= merged_range.max_col:  # Si la première colonne est dans la plage
            merged_rows_doc1.extend(
                range(merged_range.min_row - 1,
                      merged_range.max_row))
    dict_traduction = {}
    merged_rows_doc2 = []
    for merged_range in merged_cells_doc2:
        if merged_range.min_col <= 1 <= merged_range.max_col or merged_range.min_col <= 2 <= merged_range.max_col:
            merged_rows_doc2.extend(
                range(merged_range.min_row - 1,
                      merged_range.max_row))
    for index, row in doc2.iterrows():
        if (index not in merged_rows_doc2 and
                pd.notna(row[col_mot_doc2]) and
                pd.notna(row[col_traduction_doc2]) and
                str(row[col_mot_doc2]).strip() != "" and
                str(row[
                        col_traduction_doc2]).strip() != ""):
            dict_traduction[
                str(row[
                        col_mot_doc2]).strip().lower()] = str(
                row[col_traduction_doc2]).strip()
    header_row = 1  # Généralement, l'en-tête est dans la première ligne
    colonne_index = None
    for col in range(1, ws1.max_column + 1):
        if ws1.cell(row=header_row,
                    column=col).value == colonne_sortie:
            colonne_index = col
            break
    if colonne_index is None:
        colonne_index = ws1.max_column + 1
        ws1.cell(row=header_row,
                 column=colonne_index).value = colonne_sortie
    mots_traduits = 0
    for index, row in doc1.iterrows():
        excel_row = index + 2  # +2 car pandas index commence à 0 et Excel compte l'en-tête

        if (index not in merged_rows_doc1 and
                pd.notna(row[col_mot_doc1]) and
                str(row[col_mot_doc1]).strip() != ""):
            mot = str(row[col_mot_doc1]).strip()
            mot_lower = mot.lower()  # Convertir en minuscules pour la recherche
            if mot_lower in dict_traduction:
                # Si oui, récupérer sa traduction et l'écrire directement dans la cellule
                ws1.cell(row=excel_row,
                         column=colonne_index).value = \
                dict_traduction[mot_lower]
                mots_traduits += 1
    wb1.save(fichier_doc1)
    print(
        f"Traduction terminée. {mots_traduits} mots traduits.")
    print(
        f"Les modifications ont été sauvegardées dans le fichier '{fichier_doc1}' en préservant le formatage original.")
    print(
        f"La recherche a été effectuée sans distinction entre majuscules et minuscules.")

# Exemple d'utilisation
if __name__ == "__main__":
    # Remplacer avec les chemins de vos fichiers
    fichier_document1 = "Yambeta_DICTIONARY_KAAMA_ENTERPRISE.xlsx"
    fichier_document2 = "EN_FR_Fulfulde_DC_DICTIONARY-TO CLEAN.xlsx"
    colonne_pour_traduction = "Gohmala"

    traduire_mots(fichier_document1, fichier_document2,
                  colonne_pour_traduction)
    print(
        f"Les cellules vides et les cellules fusionnées ont été ignorées et préservées.")